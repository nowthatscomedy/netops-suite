from __future__ import annotations

import io
import logging
import os
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import PatternFill

from core.i18n import t
from core.settings import canonicalize_column_name

try:
    import msoffcrypto
except ImportError:
    msoffcrypto = None

logger = logging.getLogger(__name__)


def read_command_file(filepath: str) -> list[str]:
    path = Path(filepath)
    suffix = path.suffix.lower()
    commands: list[str] = []

    try:
        if suffix in (".xlsx", ".xls", ".xlsm"):
            df = pd.read_excel(filepath, header=None)
            if df.empty:
                return []
            for value in df.iloc[:, 0].tolist():
                if pd.isna(value):
                    continue
                line = str(value).strip()
                if line:
                    commands.append(line)
        elif suffix == ".txt":
            with open(filepath, "r", encoding="utf-8") as file:
                for line in file:
                    cleaned = line.strip()
                    if cleaned:
                        commands.append(cleaned)
        else:
            raise ValueError(t("file_handler.error.unsupported_extension", suffix=suffix))
    except Exception as exc:
        logger.error(
            t(
                "file_handler.error.command_file_read_failed",
                filepath=filepath,
                error=exc,
            ),
        )
        raise

    return commands


def read_excel_file(filepath: str, password: str | None = None) -> pd.DataFrame:
    try:
        if password:
            if msoffcrypto is None:
                raise ImportError(t("file_handler.error.encrypted_excel_dependency"))

            decrypted_file = io.BytesIO()
            with open(filepath, "rb") as file:
                office_file = msoffcrypto.OfficeFile(file)
                office_file.load_key(password=password)
                office_file.decrypt(decrypted_file)

            df = pd.read_excel(decrypted_file)
            logger.info(t("file_handler.info.encrypted_excel_decrypted", filepath=filepath))
        else:
            df = pd.read_excel(filepath)
            logger.info(t("file_handler.info.selected_file", filepath=filepath))
        return df
    except Exception as exc:
        logger.error(
            t("file_handler.error.excel_read_failed", filepath=filepath, error=exc),
        )
        raise


def save_results_to_excel(
    results: list[dict[str, Any]],
    output_filepath: str,
    column_order: list[str] | None = None,
    column_aliases: dict[str, str] | None = None,
) -> None:
    try:
        logger.info(t("file_handler.info.result_save_started", filepath=output_filepath))

        output_dir = os.path.dirname(output_filepath)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)

        status_col = t("excel.columns.connection_status")
        error_col = t("excel.columns.error_message")
        status_success = t("excel.status.success")
        status_failed = t("excel.status.failed")
        sheet_name = t("excel.sheet.inspection_results")

        aliases = dict(column_aliases or {})
        processed_results: list[dict[str, Any]] = []
        for result in results:
            row: dict[str, Any] = {
                "ip": result.get("ip"),
                "vendor": result.get("vendor"),
                "os": result.get("os"),
                status_col: status_success
                if result.get("status") == "success"
                else status_failed,
                error_col: result.get("error_message", ""),
            }
            inspection_results = result.get("inspection_results")
            if isinstance(inspection_results, dict):
                for key, value in inspection_results.items():
                    if key.startswith("error_") or key in {"error", "backup_error", "backup_file"}:
                        continue
                    canonical_key = canonicalize_column_name(key, aliases)
                    if not canonical_key:
                        continue
                    if (
                        canonical_key in row
                        and row[canonical_key] not in (None, "")
                        and value not in (None, "")
                        and str(row[canonical_key]) != str(value)
                    ):
                        row[canonical_key] = f"{row[canonical_key]}, {value}"
                    else:
                        row[canonical_key] = value

            processed_results.append(row)

        if not processed_results:
            logger.warning(t("file_handler.warning.no_results"))
            return

        df = pd.DataFrame(processed_results)
        base_cols = ["ip", "vendor", "os", status_col, error_col]
        if column_order:
            ordered_inspection_cols = [
                col for col in column_order if col in df.columns and col not in base_cols
            ]
            remaining_cols = [
                col
                for col in df.columns
                if col not in base_cols and col not in ordered_inspection_cols
            ]
            df = df[base_cols + ordered_inspection_cols + remaining_cols]
        else:
            other_cols = [col for col in df.columns if col not in base_cols]
            df = df[base_cols + other_cols]

        with pd.ExcelWriter(output_filepath, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            worksheet = writer.sheets[sheet_name]

            light_red_fill = PatternFill(
                start_color="FFFFC7CE",
                end_color="FFFFC7CE",
                fill_type="solid",
            )

            for index, row in df.iterrows():
                if row[status_col] != status_failed:
                    continue
                for col_idx in range(1, len(df.columns) + 1):
                    worksheet.cell(row=index + 2, column=col_idx).fill = light_red_fill

            for column_cells in worksheet.columns:
                try:
                    length = max(len(str(cell.value)) for cell in column_cells if cell.value)
                    worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2
                except (TypeError, ValueError):
                    continue

        logger.info(t("file_handler.info.result_saved", filepath=output_filepath))
    except Exception as exc:
        logger.error(t("file_handler.error.excel_save_failed", error=exc))
        raise
