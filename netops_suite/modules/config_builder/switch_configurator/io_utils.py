from __future__ import annotations

import csv
from io import BytesIO, StringIO
from pathlib import Path
from typing import Iterable

import yaml
from openpyxl import load_workbook

from .models import (
    AUTO_INCREMENT_MODES,
    AUTO_INCREMENT_NONE,
    BlockSpec,
    DeviceRecord,
    Profile,
    ValidationIssue,
    VariableSpec,
)


SUPPORTED_PROFILE_EXTENSIONS = {".yaml", ".yml"}


def parse_profile_yaml(text: str, source: str) -> Profile:
    raw = yaml.safe_load(text) or {}
    if not isinstance(raw, dict):
        raise ValueError("프로파일 YAML 최상위 구조는 object(dict)여야 합니다.")

    try:
        profile_id = str(raw["id"]).strip()
        vendor = str(raw["vendor"]).strip()
        model = str(raw["model"]).strip()
        firmware = str(raw["firmware"]).strip()
    except KeyError as exc:
        raise ValueError(f"필수 필드가 없습니다: {exc.args[0]}") from exc

    variables_raw = raw.get("variables", {})
    if not isinstance(variables_raw, dict):
        raise ValueError("variables는 key/value 구조여야 합니다.")

    blocks_raw = raw.get("blocks", [])
    if not isinstance(blocks_raw, list):
        raise ValueError("blocks는 리스트여야 합니다.")

    variables: dict[str, VariableSpec] = {}
    for variable_name, config in variables_raw.items():
        if not isinstance(config, dict):
            raise ValueError(f"variables.{variable_name} 항목은 object(dict)여야 합니다.")
        variable_key = str(variable_name).strip()
        auto_increment = str(config.get("auto_increment", AUTO_INCREMENT_NONE)).strip().lower() or AUTO_INCREMENT_NONE
        if auto_increment not in AUTO_INCREMENT_MODES:
            raise ValueError(
                f"variables.{variable_name}.auto_increment는 none, suffix_number, ipv4 중 하나여야 합니다."
            )
        variables[variable_key] = VariableSpec(
            name=variable_key,
            required=bool(config.get("required", False)),
            type=str(config.get("type", "string")).strip().lower(),
            default=config.get("default"),
            description=str(config.get("description", "")).strip(),
            description_ko=str(config.get("description_ko", "")).strip(),
            auto_increment=auto_increment,
        )

    blocks: list[BlockSpec] = []
    for index, block_raw in enumerate(blocks_raw, start=1):
        if not isinstance(block_raw, dict):
            raise ValueError(f"blocks[{index}] 항목은 object(dict)여야 합니다.")
        name = str(block_raw.get("name", f"block_{index}")).strip() or f"block_{index}"
        lines_raw = block_raw.get("lines", [])
        if not isinstance(lines_raw, list) or not all(
            isinstance(line, str) for line in lines_raw
        ):
            raise ValueError(f"blocks[{index}].lines는 문자열 리스트여야 합니다.")
        blocks.append(
            BlockSpec(
                name=name,
                lines=list(lines_raw),
            )
        )

    return Profile(
        id=profile_id,
        vendor=vendor,
        model=model,
        firmware=firmware,
        description=str(raw.get("description", "")).strip(),
        description_ko=str(raw.get("description_ko", "")).strip(),
        variables=variables,
        blocks=blocks,
        source=source,
    )


def load_profiles_from_directory(directory: str | Path) -> tuple[dict[str, Profile], list[ValidationIssue]]:
    base_path = Path(directory)
    profiles: dict[str, Profile] = {}
    issues: list[ValidationIssue] = []

    if not base_path.exists():
        issues.append(
            ValidationIssue(
                level="error",
                scope="profile",
                message="프로파일 디렉터리가 존재하지 않습니다.",
                source=str(base_path),
            )
        )
        return profiles, issues

    for path in sorted(base_path.iterdir()):
        if path.suffix.lower() not in SUPPORTED_PROFILE_EXTENSIONS:
            continue
        try:
            profile = parse_profile_yaml(path.read_text(encoding="utf-8"), str(path))
        except Exception as exc:
            issues.append(
                ValidationIssue(
                    level="error",
                    scope="profile",
                    message=str(exc),
                    source=str(path),
                )
            )
            continue

        if profile.id in profiles:
            issues.append(
                ValidationIssue(
                    level="error",
                    scope="profile",
                    message="중복된 profile id 입니다.",
                    source=str(path),
                    profile_id=profile.id,
                )
            )
            continue
        profiles[profile.id] = profile

    return profiles, issues


def load_profiles_from_uploads(uploaded_files: Iterable[object]) -> tuple[dict[str, Profile], list[ValidationIssue]]:
    profiles: dict[str, Profile] = {}
    issues: list[ValidationIssue] = []

    for uploaded_file in uploaded_files:
        name = getattr(uploaded_file, "name", "")
        if Path(name).suffix.lower() not in SUPPORTED_PROFILE_EXTENSIONS:
            issues.append(
                ValidationIssue(
                    level="error",
                    scope="profile",
                    message="지원하지 않는 프로파일 파일 형식입니다.",
                    source=name,
                )
            )
            continue

        try:
            content = uploaded_file.getvalue().decode("utf-8-sig")
            profile = parse_profile_yaml(content, name)
        except Exception as exc:
            issues.append(
                ValidationIssue(
                    level="error",
                    scope="profile",
                    message=str(exc),
                    source=name,
                )
            )
            continue

        if profile.id in profiles:
            issues.append(
                ValidationIssue(
                    level="error",
                    scope="profile",
                    message="중복된 profile id 입니다.",
                    source=name,
                    profile_id=profile.id,
                )
            )
            continue
        profiles[profile.id] = profile

    return profiles, issues


def load_device_records_from_path(path: str | Path) -> list[DeviceRecord]:
    path_obj = Path(path)
    suffix = path_obj.suffix.lower()
    if suffix == ".csv":
        return _load_device_records_from_csv_text(path_obj.read_text(encoding="utf-8-sig"))
    if suffix == ".xlsx":
        return _load_device_records_from_xlsx_bytes(path_obj.read_bytes())
    raise ValueError("지원하지 않는 장비 데이터 파일 형식입니다.")


def load_device_records_from_bytes(file_name: str, content: bytes) -> list[DeviceRecord]:
    suffix = Path(file_name).suffix.lower()
    if suffix == ".csv":
        return _load_device_records_from_csv_text(content.decode("utf-8-sig"))
    if suffix == ".xlsx":
        return _load_device_records_from_xlsx_bytes(content)
    raise ValueError("지원하지 않는 장비 데이터 파일 형식입니다.")


def _load_device_records_from_csv_text(text: str) -> list[DeviceRecord]:
    reader = csv.DictReader(StringIO(text))
    if not reader.fieldnames:
        raise ValueError("CSV 헤더를 찾을 수 없습니다.")
    records: list[DeviceRecord] = []
    for index, row in enumerate(reader, start=2):
        cleaned = _clean_row(row)
        if not any(cleaned.values()):
            continue
        records.append(DeviceRecord(row_number=index, values=cleaned))
    return records


def _load_device_records_from_xlsx_bytes(content: bytes) -> list[DeviceRecord]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("엑셀 파일이 비어 있습니다.")

    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    if not any(headers):
        raise ValueError("엑셀 헤더를 찾을 수 없습니다.")

    records: list[DeviceRecord] = []
    for row_number, values in enumerate(rows[1:], start=2):
        row = {
            headers[index]: values[index] if index < len(values) else ""
            for index in range(len(headers))
            if headers[index]
        }
        cleaned = _clean_row(row)
        if not any(cleaned.values()):
            continue
        records.append(DeviceRecord(row_number=row_number, values=cleaned))
    return records


def _clean_row(row: dict[str, object]) -> dict[str, str]:
    cleaned: dict[str, str] = {}
    for key, value in row.items():
        header = str(key).strip()
        if not header:
            continue
        cleaned[header] = "" if value is None else str(value).strip()
    return cleaned
