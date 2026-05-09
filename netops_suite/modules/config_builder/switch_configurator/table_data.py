from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from .models import DeviceRecord, Profile


@dataclass(slots=True)
class DeviceTable:
    path: Path | None
    headers: list[str]
    rows: list[dict[str, str]]


def load_device_table_from_path(path: str | Path) -> DeviceTable:
    path_obj = Path(path)
    suffix = path_obj.suffix.lower()
    if suffix == ".csv":
        headers, rows = _load_device_table_from_csv(path_obj)
    elif suffix == ".xlsx":
        headers, rows = _load_device_table_from_xlsx(path_obj)
    else:
        raise ValueError("지원하지 않는 장비 데이터 파일 형식입니다.")

    return DeviceTable(path=path_obj, headers=headers, rows=rows)


def save_device_table_to_path(
    path: str | Path,
    headers: list[str],
    rows: list[dict[str, Any]],
) -> None:
    path_obj = Path(path)
    normalized_headers = [str(header).strip() for header in headers if str(header).strip()]
    normalized_rows = [_normalize_table_row(row, normalized_headers) for row in rows]

    if path_obj.suffix.lower() == ".csv":
        _save_device_table_to_csv(path_obj, normalized_headers, normalized_rows)
        return
    if path_obj.suffix.lower() == ".xlsx":
        _save_device_table_to_xlsx(path_obj, normalized_headers, normalized_rows)
        return
    raise ValueError("지원하지 않는 장비 데이터 파일 형식입니다.")


def headers_for_profiles(
    existing_headers: list[str],
    profiles: dict[str, Profile],
) -> list[str]:
    headers = [str(header).strip() for header in existing_headers if str(header).strip()]
    seen = set(headers)
    if "profile_id" not in seen:
        headers.append("profile_id")
        seen.add("profile_id")
    for profile in profiles.values():
        for variable_name in profile.variables:
            if variable_name not in seen:
                headers.append(variable_name)
                seen.add(variable_name)
    return headers


def ensure_headers_for_profile(
    headers: list[str],
    profile: Profile | None,
) -> list[str]:
    normalized_headers = [str(header).strip() for header in headers if str(header).strip()]
    seen = set(normalized_headers)
    if "profile_id" not in seen:
        normalized_headers.append("profile_id")
        seen.add("profile_id")
    if profile:
        for variable_name in profile.variables:
            if variable_name not in seen:
                normalized_headers.append(variable_name)
                seen.add(variable_name)
    return normalized_headers


def make_blank_table_row(
    headers: list[str],
    profile: Profile | None = None,
) -> dict[str, str]:
    row = {header: "" for header in headers}
    if profile:
        row["profile_id"] = profile.id
        for variable_name in profile.variables:
            row.setdefault(variable_name, "")
    return row


def build_records_from_table(headers: list[str], rows: list[dict[str, Any]]) -> list[DeviceRecord]:
    normalized_headers = [str(header).strip() for header in headers if str(header).strip()]
    records: list[DeviceRecord] = []
    for row_number, row in enumerate(rows, start=2):
        cleaned = _normalize_table_row(row, normalized_headers)
        if not any(value for key, value in cleaned.items() if key != "profile_id"):
            continue
        records.append(DeviceRecord(row_number=row_number, values=cleaned))
    return records


def _load_device_table_from_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames:
            raise ValueError("CSV 헤더를 찾을 수 없습니다.")
        headers = [str(header).strip() for header in reader.fieldnames if str(header).strip()]
        rows: list[dict[str, str]] = []
        for raw_row in reader:
            row = _normalize_table_row(raw_row, headers)
            if not any(row.values()):
                continue
            rows.append(row)
    return headers, rows


def _load_device_table_from_xlsx(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    workbook = load_workbook(path, data_only=False)
    sheet = workbook.active
    values = list(sheet.iter_rows(values_only=True))
    if not values:
        raise ValueError("엑셀 파일이 비어 있습니다.")

    headers = [
        str(value).strip()
        for value in values[0]
        if value is not None and str(value).strip()
    ]
    if not headers:
        raise ValueError("엑셀 헤더를 찾을 수 없습니다.")

    rows: list[dict[str, str]] = []
    for raw_values in values[1:]:
        row = {
            headers[index]: "" if index >= len(raw_values) or raw_values[index] is None else str(raw_values[index]).strip()
            for index in range(len(headers))
        }
        if not any(row.values()):
            continue
        rows.append(row)
    return headers, rows


def _save_device_table_to_csv(
    path: Path,
    headers: list[str],
    rows: list[dict[str, str]],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers, lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow({header: row.get(header, "") for header in headers})


def _save_device_table_to_xlsx(
    path: Path,
    headers: list[str],
    rows: list[dict[str, str]],
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        workbook = load_workbook(path)
    else:
        workbook = Workbook()
    sheet = workbook.active
    if sheet.max_row:
        sheet.delete_rows(1, sheet.max_row)

    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    workbook.save(path)


def _normalize_table_row(
    row: dict[str, Any],
    headers: list[str],
) -> dict[str, str]:
    normalized: dict[str, str] = {}
    for header in headers:
        value = row.get(header, "")
        normalized[header] = "" if value is None else str(value).strip()
    return normalized
